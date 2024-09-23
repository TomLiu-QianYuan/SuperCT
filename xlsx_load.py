# 这是一个示例 Python 脚本。
from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def extract_and_create_file(dict_wrong, dict_correct: dict):
    # 该函数用于处理错误单词字典和正确单词字典并生成一个类文件(xlsx格式).
    # 使用该函数需导入以下四个模块:import openpyxl; from openpyxl.utils import get_column_letter;
    # from openpyxl.styles import Font; from io import BytesIO
    # Author:Sword
    wb = openpyxl.Workbook()
    sheet = wb.active
    counter = 0
    sheet.title = "Your Word List"
    desired_width = 70
    sheet.column_dimensions[get_column_letter(1)].width = desired_width
    for key in dict_wrong.keys():
        a = sheet.cell(counter + 1, 1, key)
        b = sheet.cell(counter + 1, 2, dict_wrong[key])
        a.font = Font(color="FF0000")
        b.font = Font(color="FF0000")
        counter += 1
    for key in dict_correct.keys():
        a = sheet.cell(counter + 1, 1, key)
        b = sheet.cell(counter + 1, 2, dict_correct[key])
        a.font = Font(color="6DB33F")
        b.font = Font(color="6DB33F")
        counter += 1
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


if __name__ == '__main__':
    # 你可以根据需要调整这个值
    test_dict_wrong = {'使人或物保持直立和稳定的重量分布均匀。': 'Balance', '突然的和引人注目的。': 'Dramatic'}
    test_dict_correct = {'使(某物)适应新的用途或目的；修改。': 'Adapt', '对象或点相对于海平面或地面的高度。': 'Altitude',
                         '人或动物的身体在颈部和腹部之间的前表面。': 'Chest',
                         '从较高的位置向较低的位置移动，通常是迅速且无法控制的。': 'Fall',
                         '某物的独特属性或方面。': 'Feature', '紧紧抓住并保持；紧握。': 'Grip',
                         '在牛、羊、山羊、长颈鹿等头上发现的硬质永久性突出物，通常是曲折和尖锐的，成对出现。': 'Horn',
                         '倾斜或斜坡。': 'Incline', '某事的例子或单一事件。': 'Instance',
                         '向听众（尤指在大学中）进行的教育性讲话。': 'Lecture',
                         '地球表面突然从周围平地升起的大自然高地；大陡峭的山丘。': 'Mountain',
                         '有许多山的。': 'Mountainous',
                         '人或动物体内能够收缩，产生运动或保持身体部位位置的纤维状组织束或带。': 'Muscle',
                         '一块厚厚的软物，通常用来保护或塑造某物，或吸收某物。': 'Pad',
                         '某人或某物在特定时刻的位置。': 'Position', '为特定目的所需要。': 'Require',
                         '由岩石或岩石组成的；表面粗糙不平。': 'Rocky',
                         '在保持与其连续接触的同时，沿着表面平滑地移动。': 'Slide',
                         '滑，意外滑动一段短距离，通常是失去平衡或立足点。': 'Slip',
                         '（指表面或物体）滑，湿，冰凉或黏滑到导致滑动或失去抓力的程度。': 'Slippery',
                         '一个表面的一端或一侧比另一侧的水平线高；上升或下降的表面。': 'Slope',
                         '身体强壮的品质或状态。': 'Strength', '某物的外部部分或最上层。': 'Surface',
                         '牢固地固定或紧固在位；不易移动或松动。': 'Tight', '人脚的末 端的任何一只手指。': 'Toe',
                         '与水平面成直角；在一个方向上，或有一个对齐方式，使得顶部直接在底部之上。': 'Vertical'}
    with open("/Users/wujijianke/Desktop/user.xlsx", "wb") as f:
        f.write(extract_and_create_file(test_dict_wrong, test_dict_correct).read())
